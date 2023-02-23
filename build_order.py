#! /usr/bin/env python3


import argparse
import collections
import datetime

import numpy as np
import openpyxl

import order_optimization as oo

ORDER_SIZE = 35
PSEUDOCOUNT = 35

SIM_SIZE = 10_000

# Initialize NumPy random number generator
gen = np.random.default_rng(seed=42)

parser = argparse.ArgumentParser(
    description='Compute an optimal T-shirt order.')
parser.add_argument('inventory_filename',
    help='filename of Excel spreadsheet with inventory')
parser.add_argument('-m', '--method',
    choices=['heuristic', 'worst_case'], default='heuristic',
    help='how to compute the optimal order')
parser.add_argument('-o', '--output',
    choices=['console', 'hypothetical', 'final'], default='console',
    help='where to write down the optimal order')
args = parser.parse_args()

# Grab counts of lifetime T-shirts received and lifetime T-shirts queued from
# the inventory spreadsheet
wb = openpyxl.load_workbook(filename=args.inventory_filename, data_only=True)
assert wb['inventory']['A2'].value == 'Lifetime received'
assert wb['inventory']['A3'].value == 'Lifetime queued'

lifetime_received = collections.OrderedDict()
lifetime_queued = collections.OrderedDict()
logical_inventory = collections.OrderedDict()  # Negatives denote backorders
gendered_sizes = []

for column in wb['inventory'].iter_cols():
    header_val = column[0].value
    if header_val is not None and header_val != 'totals':
        lifetime_received[header_val] = column[1].value
        lifetime_queued[header_val] = column[2].value
        logical_inventory[header_val] = column[6].value
        gendered_sizes.append(header_val)
wb.close()

# We define a weakly informative Dirichlet prior from industry knowledge:
industry_knowledge = {'XS': 0.01,
                      'S': 0.07,
                      'M': 0.28,
                      'L': 0.30,
                      'XL': 0.20,
                      '2XL': 0.12,
                      '3XL': 0.02}

# A bit of tidying to deal with the real world: the vendor does not sell
# women's XS or 3XL shirts. There are a couple reasonable, arbitrary ways to
# deal with this in conjunction with a 50:50 prior distribution of men vs
# women. One would be to allocate the other sizes evenly between men and women,
# allocating all 1% of XS and all 2% of 3XL to men. Another would be to
# re-normalize the distribution of women's shirts, such that (e.g.) 0.07 / (1 -
# 0.01 - 0.02) of women's shirts are size WS. Note these are very close.
#
# Doing the first thing, which makes the prior histogram 51.5% men and 48.5%
# women:
prior_size_hist = collections.OrderedDict()
for gendered_size in gendered_sizes:
    size = gendered_size[1:]
    if f'M{size}' in lifetime_received and \
            f'W{size}' in lifetime_received:
        prior_size_hist[gendered_size] = industry_knowledge[size] / 2.0
    else:
        prior_size_hist[gendered_size] = industry_knowledge[size]

alpha_prior = np.array([1.0 + PSEUDOCOUNT * val
    for val in prior_size_hist.values()])

prior_samples = gen.dirichlet(alpha_prior, size=SIM_SIZE)

# Diagnostic for choosing a reasonable pseudocount for the prior: How many MXS
# shirts were ordered? 1% of 35 shirts is 0.35 shirts on average. Ordering more
# than 3 of these, a priori, should happen pretty seldom.
#
# mxs_probs = prior_samples[:, 0]
# more_than_three = sum(gen.binomial(ORDER_SIZE, prob) > 3 for prob in mxs_probs)
# print(more_than_three / SIM_SIZE)
#
# Some prior predictive simulation results:
# PSEUDOCOUNT = 0 => 32% of orders have >3 MXS shirts
# PSEUDOCOUNT = 20 => 9.2% of orders have >3 MXS shirts
# PSEUDOCOUNT = 35 => 5.0% of orders have >3 MXS shirts
# PSEUDOCOUNT = 50 => 3.3% of orders have >3 MXS shirts
#
# From this, PSEUDOCOUNT = 35 seems pretty reasonable: flexible enough to allow
# for startlingly lopsided orders, but not flexible enough that they happen all
# the time.

# Construct posterior from lifetime queued
counts = np.array(list(lifetime_queued.values()))
alpha_posterior = alpha_prior + counts

# Now sample distributions from the posterior:
posterior_samples = gen.dirichlet(alpha_posterior, size=SIM_SIZE)

# At most we'd need to simulate logical_inventory + ORDER_SIZE samples from
# each draw of the Dirichlet distribution (i.e., one could never do better than
# perfect efficiency where you fill existing backorders, then use every single
# T-shirt already in stock and every T-shirt in the order we're planning here,
# getting to a completely empty inventory before the following re-order).
inv_arr = np.array(list(logical_inventory.values()))

# Build the library of order streams.
stream_length = sum(inv_arr) + ORDER_SIZE
order_streams = np.zeros((posterior_samples.shape[0], stream_length),
        dtype='int')
for i in range(order_streams.shape[0]):
    order_streams[i, :] = gen.choice(inv_arr.shape[0],
            size=stream_length,
            p=posterior_samples[i])

# Compute optimal order given SIM_SIZE streams of future orders and the current
# logical inventory
if args.method == 'heuristic':
    optimal_order = oo.heuristic(inv_arr, order_streams, ORDER_SIZE)
elif args.method == 'worst_case':
    optimal_order = oo.worst_case(inv_arr, order_streams, ORDER_SIZE)

# Output to console or to the inventory workbook itself, either as a
# hypothetical order at the bottom of the inventory sheet (to think about) or
# as a new line in the incoming sheet (as part of sending out a new order)
if args.output == 'console':
    print('Optimal order:')
    for i, gendered_size in enumerate(gendered_sizes):
        print('{:4s}: {:d}'.format(gendered_size, optimal_order[i]))
elif args.output == 'hypothetical':
    # With formulas, so we can save the formulas
    wb = openpyxl.load_workbook(filename=args.inventory_filename)
    assert wb['inventory']['A17'].value == 'Hypothetical order'
    for column in wb['inventory'].iter_cols():
        header_val = column[0].value
        if header_val is not None and header_val != 'totals':
            i = gendered_sizes.index(header_val)
            column[16].value = optimal_order[i]
    wb.save(args.inventory_filename)
elif args.output == 'final':
    # With formulas, so we can save the formulas
    wb = openpyxl.load_workbook(filename=args.inventory_filename)
    ws = wb['incoming']
    row_to_write = ws.max_row + 1

    # Today's date for the reorder line
    assert ws.cell(row=1, column=1).value == 'date'
    new_date_cell = ws.cell(row=row_to_write, column=1)
    new_date_cell.value = datetime.date.today()
    new_date_cell.number_format = 'm/d/yy'

    # Order quantities
    for j in range(1, ws.max_column):
        col_to_write = j + 1
        new_qty_cell = ws.cell(row=row_to_write, column=col_to_write)
        assert new_qty_cell.value is None
        i = gendered_sizes.index(ws.cell(row=1, column=col_to_write).value)
        ws.cell(row=row_to_write, column=col_to_write).value = \
                optimal_order[i]

    wb.save(args.inventory_filename)
